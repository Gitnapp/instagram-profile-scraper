import pandas as pd
from instagrapi import Client
import time
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# 设置日志记录
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 读取Excel文件中的Profile URL
def read_profile_urls(file_path, start_row, end_row):
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        return df['ProfileURL'][start_row-1:end_row].tolist()
    except Exception as e:
        logging.error(f"Error reading profile URLs from {file_path}: {e}")
        return []

# 提取用户名
def extract_username(profile_url):
    # 去掉末尾的斜杠
    if profile_url.endswith('/'):
        profile_url = profile_url[:-1]
    # 提取用户名
    username = profile_url.split('/')[-1]
    return username

# 抓取Instagram Profile数据
def scrape_instagram_profile(profile_url, session_cookie):
    client = Client()
    client.login_by_sessionid(session_cookie)
    
    username = extract_username(profile_url)
        
    # 检查 username 是否为空
    if not username:
        logging.error(f"Username is empty for profile URL: {profile_url}")
        return {'profileUrl': profile_url, 'error': 'Username is empty'}
    
    try:
        user_info = client.user_info_by_username_v1(username)
        logging.info(f"scraping {username}, {profile_url}, 200")
    except Exception as e:
        logging.error(f"scraping {username}, {profile_url}, 404: {e}")
        return {'profileUrl': profile_url, 'error': str(e)}
    
    if user_info:
        return {
            'profileUrl': profile_url,
            'profileName': user_info.username,
            'instagramID': user_info.pk,
            'fullName': user_info.full_name,
            'postsCount': user_info.media_count,
            'followersCount': user_info.follower_count,
            'followingCount': user_info.following_count,
            'bio': user_info.biography,
            'verified': user_info.is_verified,
            'private': user_info.is_private
        }
    else:
        logging.warning(f"No user info found for {profile_url}")
        return None

# 保存数据到Excel文件
def save_to_excel(data, output_file):
    df = pd.DataFrame(data)
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file, engine='openpyxl')
            df = pd.concat([existing_df, df], ignore_index=True)
        except Exception as e:
            logging.error(f"Error reading existing file {output_file}: {e}")
            return
    else:
        logging.info(f"Creating new file: {output_file}")
    
    # 标记错误行
    try:
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for idx, row in df.iterrows():
            if 'error' in row:
                for cell in worksheet[idx + 2]:  # +2 because of 0-based index and header row
                    cell.fill = red_fill
        
        writer.close()
        logging.info(f"Data saved to {output_file}")
    except Exception as e:
        logging.error(f"Error saving data to {output_file}: {e}")

# 主函数
def main():
    input_file = 'profile_urls.xlsx'  # 输入文件路径
    output_file = 'instagram_profiles.xlsx'  # 输出文件路径
    session_cookie = 'YOUR_INSTAGRAM_SESSION_COOKIE'  # 替换为你的Instagram Session Cookie
    
    # 获取用户输入的行数范围
    row_range = input("Enter row range to read (start-end): ")
    start_row, end_row = map(int, row_range.split('-'))
    
    # 获取用户输入的爬取用户数、每个请求间隔时间和每组暂停时间
    group_info = input("Enter number of users per group, interval between requests (seconds), and pause time after each group (seconds) (users,interval,pause_time): ")
    users_per_group, request_interval, group_pause_time = map(int, group_info.split(','))
    
    profile_urls = read_profile_urls(input_file, start_row, end_row)
    scraped_data = []
    count = 0
    
    logging.info(f"Reading profile URLs from {input_file} from row {start_row} to {end_row}")
    
    for profile_url in profile_urls:
        profile_data = None
        retries = 3
        for attempt in range(retries):
            try:
                profile_data = scrape_instagram_profile(profile_url, session_cookie)
                break
            except Exception as e:
                logging.error(f"Attempt {attempt + 1} failed for {profile_url}: {e}")
                time.sleep(request_interval)
        
        if profile_data:
            scraped_data.append(profile_data)
            count += 1
        
        if count % users_per_group == 0:
            logging.info(f"Scraped {count} users. Pausing for {group_pause_time} seconds...")
            save_to_excel(scraped_data, output_file)
            scraped_data = []
            time.sleep(group_pause_time)
        else:
            logging.info(f"Waiting for {request_interval} seconds before next request...")
            time.sleep(request_interval)
    
    if scraped_data:
        save_to_excel(scraped_data, output_file)
    
    logging.info(f"Data saved to {output_file}")

if __name__ == "__main__":
    main()
